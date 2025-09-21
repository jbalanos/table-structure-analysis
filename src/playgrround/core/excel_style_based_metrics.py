from __future__ import annotations
from pathlib import Path
from typing import Optional, Tuple, Iterable, List

import numpy as np
import pandas as pd

# We import openpyxl lazily inside the main function to keep this module import-light.


# ---------------------------
# Private helpers (style predicates & utilities)
# ---------------------------

_THICK_STYLES = {"medium", "thick", "double"}

def _is_nonwhite_fill(cell) -> bool:
    """
    Return True if the cell has a meaningful (non-white) background fill.
    We treat any explicit nonempty fill color as non-white; theme/indexed counts as non-default.
    """
    f = getattr(cell, "fill", None)
    if f is None or f.fill_type in (None, "none"):
        return False
    c = getattr(f, "fgColor", None)
    if not c:
        return False
    t = getattr(c, "type", None)
    if t == "rgb":
        rgb = (c.rgb or "").upper()
        # White is usually FFFFFF with alpha FF prefix; also guard against empty.
        return rgb not in ("FFFFFFFF", "FF000000", "", None)
    # If theme/indexed is present, consider it non-default (useful as a weak signal).
    return True


def _is_nondefault_font_color(cell) -> bool:
    """
    Return True if the font color is not the default black.
    If color is theme/indexed, treat as non-default (weak signal).
    """
    font = getattr(cell, "font", None)
    if not font:
        return False
    col = getattr(font, "color", None)
    if not col:
        return False
    t = getattr(col, "type", None)
    if t == "rgb":
        rgb = (col.rgb or "").upper()
        return rgb not in ("FF000000", "", None)
    return True


def _has_thick_top(cell) -> bool:
    """Return True if the cell's top border is medium/thick/double."""
    top = getattr(getattr(cell, "border", None), "top", None)
    style = getattr(top, "style", None)
    return style in _THICK_STYLES


def _has_thick_bottom(cell) -> bool:
    """Return True if the cell's bottom border is medium/thick/double."""
    bottom = getattr(getattr(cell, "border", None), "bottom", None)
    style = getattr(bottom, "style", None)
    return style in _THICK_STYLES


def _cell_is_nonnull(cell) -> bool:
    """Treat None or empty-string as null; everything else non-null."""
    v = getattr(cell, "value", None)
    return not (v is None or (isinstance(v, str) and v.strip() == ""))


def _cell_is_numeric(cell) -> bool:
    """
    Heuristic for numeric cells without parsing:
    - openpyxl cell.data_type == 'n' or
    - value is int/float (excluding booleans).
    """
    v = getattr(cell, "value", None)
    if isinstance(v, bool):
        return False
    if isinstance(v, (int, float)):
        # numpy floats/ints also pass as float/int
        return not (isinstance(v, float) and (np.isnan(v)))
    dt = getattr(cell, "data_type", None)
    return dt == "n"


def _right_aligned(cell) -> bool:
    """True if the cell alignment is explicitly right-aligned."""
    align = getattr(cell, "alignment", None)
    return getattr(align, "horizontal", None) == "right"


def _indent_level(cell) -> int:
    """Return alignment.indent (or 0 when missing)."""
    align = getattr(cell, "alignment", None)
    ind = getattr(align, "indent", None)
    try:
        return int(ind or 0)
    except Exception:
        return 0


def _merged_rows_mask(ws, start_row: int, end_row: int, min_col: int, max_col: int) -> np.ndarray:
    """
    Build an array of shape (end_row-start_row+1,) with 1 where the row intersects
    any merged cell range within the [min_col, max_col] window, else 0.
    """
    mask = np.zeros(end_row - start_row + 1, dtype=np.uint8)
    # openpyxl stores merged ranges as cell ranges like A1:C3
    for rng in ws.merged_cells.ranges:
        rmin, rmax = rng.min_row, rng.max_row
        cmin, cmax = rng.min_col, rng.max_col
        # Check intersection with our window
        if rmax < start_row or rmin > end_row:
            continue
        if cmax < min_col or cmin > max_col:
            continue
        # Mark all overlapping rows
        r0 = max(rmin, start_row)
        r1 = min(rmax, end_row)
        mask[(r0 - start_row):(r1 - start_row + 1)] = 1
    return mask


# ---------------------------
# Public API
# ---------------------------

def excel_style_metrics(
    xlsx_path: str | Path,
    df: pd.DataFrame,
    sheet: Optional[str | int] = None,
    start_row: int = 1,
    start_col: int = 1,
    ) -> pd.DataFrame:
    """
    Compute per-row Excel style/formatting metrics (B1–B3) aligned to `df` rows.

    Parameters
    ----------
    xlsx_path : str | Path
        Path to the Excel workbook. Supports .xlsx/.xlsm for styles.
        For .xlsb/.xls, styles are not reliably available; returns defaults.
    df : pd.DataFrame
        DataFrame whose shape defines the (rows x columns) window to examine.
        We assume the sheet area of interest starts at (start_row, start_col).
    sheet : str | int | None
        Worksheet name or 0-based index. If None, the active sheet is used.
    start_row : int
        1-based row index in the sheet where the `df`'s first row maps.
    start_col : int
        1-based column index in the sheet where the `df`'s first column maps.

    Returns
    -------
    pd.DataFrame
        One row per `df` row, indexed to `df.index`, with columns:

        B1 (Emphasis & fills)
          - bold_ratio (float)                 : fraction of non-null cells with font.bold=True
          - any_bold_row (bool)                : any bold among non-null cells
          - nonwhite_fill_ratio (float)        : fraction of non-null cells with non-white fills
          - has_nonwhite_fill (bool)           : any non-white fill among non-null cells
          - nondefault_fontcolor_ratio (float) : fraction of non-null cells with non-default font color

        B2 (Borders & merges)
          - has_thick_top_border (bool)        : any cell with top border medium/thick/double
          - has_thick_bottom_border (bool)     : any cell with bottom border medium/thick/double
          - row_has_merged (bool)              : row intersects any merged cell range (in window)

        B3 (Alignment/indent)
          - right_aligned_numeric_ratio (float): among numeric cells, fraction right-aligned
          - mean_indent_level (float)          : average indent across non-null cells

        For unsupported formats (.xlsb/.xls), all ratios are 0.0 and booleans False.

    Notes
    -----
    - This function reads the Excel *styles*; it does not rely on pandas-loaded values.
    - For performance, it scans exactly the window defined by `df.shape` starting at (start_row, start_col).
    """
    xlsx_path = Path(xlsx_path)
    n_rows, n_cols = df.shape

    # Fast fallback for formats without reliable style access
    if xlsx_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        return _style_defaults(len(df), df.index)

    try:
        from openpyxl import load_workbook
    except ImportError as e:
        # If openpyxl isn't available, degrade gracefully to defaults
        return _style_defaults(len(df), df.index)

    # Load workbook with styles available. (read_only=False to ensure style objects are present)
    wb = load_workbook(filename=str(xlsx_path), data_only=True, read_only=False)
    if sheet is None:
        ws = wb.active
    elif isinstance(sheet, int):
        ws = wb.worksheets[sheet]
    else:
        ws = wb[sheet]

    # Compute the scan window in Excel coordinates
    min_row = start_row
    max_row = start_row + n_rows - 1
    min_col = start_col
    max_col = start_col + n_cols - 1

    # Precompute merged-rows mask for the window
    merged_mask = _merged_rows_mask(ws, min_row, max_row, min_col, max_col)

    # Prepare containers
    bold_ratio = np.zeros(n_rows, dtype=float)
    any_bold = np.zeros(n_rows, dtype=bool)
    fill_ratio = np.zeros(n_rows, dtype=float)
    any_fill = np.zeros(n_rows, dtype=bool)
    fontcol_ratio = np.zeros(n_rows, dtype=float)

    thick_top = np.zeros(n_rows, dtype=bool)
    thick_bottom = np.zeros(n_rows, dtype=bool)
    row_has_merged = merged_mask.astype(bool)

    right_align_numeric_ratio = np.zeros(n_rows, dtype=float)
    mean_indent = np.zeros(n_rows, dtype=float)

    # Iterate once across the target window; compute per-row aggregates.
    # Using ws.iter_rows ensures we access style objects only once per cell.
    row_iter = ws.iter_rows(min_row=min_row, max_row=max_row,
                            min_col=min_col, max_col=max_col, values_only=False)

    for r_i, row_cells in enumerate(row_iter):
        # r_i is 0-based within our window; Excel row number is min_row + r_i
        # We compute counters among NON-NULL cells for ratios (as defined in spec).
        nn_count = 0                       # non-null cell count
        bold_cnt = 0
        fill_cnt = 0
        fontcol_cnt = 0

        thick_top_any = False
        thick_bottom_any = False

        numeric_cnt = 0
        right_aligned_numeric_cnt = 0

        indent_sum = 0

        for c in row_cells:
            nonnull = _cell_is_nonnull(c)
            if nonnull:
                nn_count += 1
                # B1: Emphasis & fills
                if getattr(getattr(c, "font", None), "bold", False):
                    bold_cnt += 1
                if _is_nonwhite_fill(c):
                    fill_cnt += 1
                if _is_nondefault_font_color(c):
                    fontcol_cnt += 1

                # B3: Alignment/indent (computed over relevant cells)
                indent_sum += _indent_level(c)

            # B2: Borders (consider any cell in the row, not only non-nulls)
            if _has_thick_top(c):
                thick_top_any = True
            if _has_thick_bottom(c):
                thick_bottom_any = True

            # B3: Right-aligned numeric ratio — among numeric cells only
            if _cell_is_numeric(c):
                numeric_cnt += 1
                if _right_aligned(c):
                    right_aligned_numeric_cnt += 1

        # Aggregate ratios and flags for this row
        bold_ratio[r_i] = (bold_cnt / nn_count) if nn_count else 0.0
        any_bold[r_i] = bold_cnt > 0

        fill_ratio[r_i] = (fill_cnt / nn_count) if nn_count else 0.0
        any_fill[r_i] = fill_cnt > 0

        fontcol_ratio[r_i] = (fontcol_cnt / nn_count) if nn_count else 0.0

        thick_top[r_i] = thick_top_any
        thick_bottom[r_i] = thick_bottom_any

        right_align_numeric_ratio[r_i] = (right_aligned_numeric_cnt / numeric_cnt) if numeric_cnt else 0.0
        mean_indent[r_i] = (indent_sum / nn_count) if nn_count else 0.0

    # Build output DataFrame aligned to df.index
    out = pd.DataFrame({
        # B1 — Emphasis & fills
        "bold_ratio": bold_ratio,
        "any_bold_row": any_bold,
        "nonwhite_fill_ratio": fill_ratio,
        "has_nonwhite_fill": any_fill,
        "nondefault_fontcolor_ratio": fontcol_ratio,

        # B2 — Borders & merges
        "has_thick_top_border": thick_top,
        "has_thick_bottom_border": thick_bottom,
        "row_has_merged": row_has_merged,

        # B3 — Alignment/indent
        "right_aligned_numeric_ratio": right_align_numeric_ratio,
        "mean_indent_level": mean_indent,
    }, index=df.index)

    # Ensure dtypes are clean
    bool_cols = ["any_bold_row", "has_nonwhite_fill", "has_thick_top_border",
                 "has_thick_bottom_border", "row_has_merged"]
    for c in bool_cols:
        out[c] = out[c].astype(bool)

    float_cols = ["bold_ratio", "nonwhite_fill_ratio", "nondefault_fontcolor_ratio",
                  "right_aligned_numeric_ratio", "mean_indent_level"]
    for c in float_cols:
        out[c] = out[c].astype(float)

    return out


def _style_defaults(n_rows: int, index: pd.Index) -> pd.DataFrame:
    """
    Return a DataFrame of default (neutral) style metrics when styles are unavailable.
    All ratios are 0.0 and flags False.
    """
    zeros = np.zeros(n_rows, dtype=float)
    falses = np.zeros(n_rows, dtype=bool)
    return pd.DataFrame({
        "bold_ratio": zeros,
        "any_bold_row": falses,
        "nonwhite_fill_ratio": zeros,
        "has_nonwhite_fill": falses,
        "nondefault_fontcolor_ratio": zeros,
        "has_thick_top_border": falses,
        "has_thick_bottom_border": falses,
        "row_has_merged": falses,
        "right_aligned_numeric_ratio": zeros,
        "mean_indent_level": zeros,
    }, index=index)
